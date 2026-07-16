"""
Servicios de exportación de la auditoría (Excel y PDF).

Separa la lógica de exportación de las vistas. No modifica la base de datos.
"""
from django.http import HttpResponse


def exportar_auditoria_excel(queryset, filename="auditoria.xlsx"):
    """
    Genera un archivo Excel con los eventos de auditoría del queryset.
    Devuelve un HttpResponse listo para descargar, o None si falta openpyxl.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    encabezados = [
        "ID", "Fecha y hora", "Usuario", "Acción", "Tabla afectada",
        "ID registro", "Descripción", "IP origen",
    ]
    ws.append(encabezados)
    relleno = PatternFill("solid", fgColor="5A3825")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno

    for a in queryset:
        usuario = a.id_usuario.nombre_completo if a.id_usuario_id else "—"
        ws.append([
            a.id_auditoria,
            a.fecha_hora.strftime("%Y-%m-%d %H:%M:%S") if a.fecha_hora else "",
            usuario,
            a.get_accion_display(),
            a.tabla_afectada,
            a.id_registro_afectado if a.id_registro_afectado is not None else "",
            (a.descripcion or "")[:500],
            a.ip_origen or "",
        ])

    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 2, 60)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def exportar_auditoria_pdf(queryset, filtros, filename="auditoria.pdf", limite=1000):
    """
    Genera un PDF con los eventos de auditoría (reutiliza el renderizador de
    apps/reportes/pdf.py, que ya tiene respaldo con ReportLab).
    """
    from django.utils import timezone

    from apps.reportes.pdf import render_pdf_bytes

    contexto = {
        "eventos": list(queryset[:limite]),
        "filtros": filtros,
        "planta": "Choco Pasión - Tingo María",
        "fecha_generacion": timezone.now(),
        "total": queryset.count(),
        "limite": limite,
    }
    pdf_bytes = render_pdf_bytes("core/auditoria_pdf.html", contexto)
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# =====================================================================
# Análisis de activos: prioridad preventiva y repuestos críticos por máquina
# (solo lectura; no modifica la base de datos)
# =====================================================================
_PESO_CRITICIDAD = {"BAJA": 1, "MEDIA": 2, "ALTA": 3, "MUY_ALTA": 4, "CRITICA": 4}
_PESO_RIESGO = {"BAJO": 1, "MEDIO": 2, "ALTO": 3, "CRITICO": 4}
_PESO_ESTADO = {
    "BUENO": 0, "OBSERVADO": 1, "CAMBIADO": 1, "DANADO": 3, "FUERA_DE_SERVICIO": 4,
}


def repuestos_criticos_maquina(maquina):
    """
    Devuelve la lista de repuestos críticos de una máquina (tabla
    maquina_repuesto) con el estado de stock calculado a partir de `repuestos`.
    Cada elemento incluye el atributo `.stock_bajo` (bool).
    """
    from apps.inventario.models import MaquinaRepuesto

    items = list(
        MaquinaRepuesto.objects.select_related("id_repuesto", "id_repuesto__id_proveedor")
        .filter(id_maquina=maquina)
        .order_by("-criticidad_repuesto")
    )
    for it in items:
        rep = it.id_repuesto
        it.stock_bajo = rep is not None and rep.stock_actual <= rep.stock_minimo
    return items


def hay_stock_bajo(items_repuestos):
    return any(getattr(it, "stock_bajo", False) for it in items_repuestos)


def evaluar_prioridad_componente(componente, fallas_del_componente, stock_bajo_asociado=False):
    """
    Calcula la prioridad de revisión preventiva de un componente combinando:
      - criticidad del componente
      - máximo nivel de riesgo de sus fallas probables
      - estado actual del componente
      - si hay repuestos asociados con stock bajo
    Devuelve un dict con score, etiqueta (ALTA/MEDIA/BAJA) y clase de color.
    """
    score = _PESO_CRITICIDAD.get(componente.criticidad_componente, 2)

    riesgo_max = 0
    for f in fallas_del_componente:
        riesgo_max = max(riesgo_max, _PESO_RIESGO.get(f.nivel_riesgo, 0))
    score += riesgo_max

    score += _PESO_ESTADO.get(componente.estado_componente, 0)
    if stock_bajo_asociado:
        score += 2

    if score >= 8:
        etiqueta, badge = "ALTA", "danger"
    elif score >= 5:
        etiqueta, badge = "MEDIA", "warning"
    else:
        etiqueta, badge = "BAJA", "success"
    return {"score": score, "etiqueta": etiqueta, "badge": badge, "riesgo_max": riesgo_max}


def analizar_activos_maquina(maquina):
    """
    Construye el análisis integral de una máquina para el detalle:
    componentes (con prioridad preventiva), fallas probables, repuestos críticos
    y las banderas de alerta. Todo en una sola función para las vistas.
    """
    componentes = list(maquina.componentes.all())
    fallas = list(maquina.fallas_probables.select_related("id_componente"))
    repuestos = repuestos_criticos_maquina(maquina)
    stock_bajo_global = hay_stock_bajo(repuestos)

    fallas_por_componente = {}
    for f in fallas:
        fallas_por_componente.setdefault(f.id_componente_id, []).append(f)

    for c in componentes:
        c.prioridad = evaluar_prioridad_componente(
            c, fallas_por_componente.get(c.pk, []), stock_bajo_global
        )

    componentes.sort(key=lambda c: c.prioridad["score"], reverse=True)

    componentes_danados = [
        c for c in componentes if c.estado_componente in ("DANADO", "FUERA_DE_SERVICIO")
    ]
    fallas_criticas = [f for f in fallas if f.nivel_riesgo == "CRITICO"]
    fallas_altas = [f for f in fallas if f.nivel_riesgo == "ALTO"]
    repuestos_bajo_stock = [it for it in repuestos if getattr(it, "stock_bajo", False)]

    return {
        "componentes": componentes,
        "fallas_probables": fallas,
        "repuestos_criticos": repuestos,
        "componentes_danados": componentes_danados,
        "fallas_criticas": fallas_criticas,
        "fallas_altas": fallas_altas,
        "fallas_riesgo_alto": fallas_criticas + fallas_altas,
        "repuestos_bajo_stock": repuestos_bajo_stock,
        "hay_stock_bajo": stock_bajo_global,
    }
